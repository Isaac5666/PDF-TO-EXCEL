from openpyxl import load_workbook


class ExcelWriter:

    def __init__(self, template_path):

        self.template_path = template_path

    def write_data(self, data):

        wb = load_workbook(
            self.template_path
        )

        ws = wb.active

        HEADER_ROW = 2

        WRITE_ROW = ws.max_row + 1

        headers = {}

        for col in range(
            1,
            ws.max_column + 1
        ):

            header = ws.cell(
                HEADER_ROW,
                col
            ).value

            if header:

                headers[
                    str(header).strip()
                ] = col

        print("\nHEADERS FOUND:\n")

        for h in headers:
            print(h)

        for key, value in data.items():

            if key in headers:

                print(
                    f"WRITING -> {key} = {value}"
                )

                ws.cell(
                    WRITE_ROW,
                    headers[key]
                ).value = value

        wb.save(
            "output/extracted.xlsx"
        )

        return WRITE_ROW